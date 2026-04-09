#!/usr/bin/env python3
"""
Local News Web Scraper
======================
A general-purpose scraper that pulls articles from local news websites
and exports them to an Excel file (.xlsx).

Usage:
    python news_scraper.py                          # Interactive mode
    python news_scraper.py -u https://example.com   # Single site
    python news_scraper.py -f sites.txt             # File with URLs (one per line)
    python news_scraper.py -u https://a.com https://b.com  # Multiple sites
    python news_scraper.py --sheet "https://docs.google.com/spreadsheets/d/SHEET_ID/edit"
    python news_scraper.py --sheet SHEET_ID --sheet-tab "My Tab" --sheet-column B
    python news_scraper.py --sheet SHEET_ID --sheet-output-column B  # column B = output file per site

Output:
    Each run produces an Excel (.xlsx) file with two sheets:
      - Articles  : one row per article (title, author, date, body, url, source)
      - Keywords  : meaningful words extracted from titles/bodies, ranked by frequency

Dependencies (required):
    pip install requests beautifulsoup4 lxml openpyxl

Dependencies (optional, for JS-rendered pages):
    pip install selenium webdriver-manager
"""

import csv
import io
import os
import re
import sys
import time
import logging
import argparse
import hashlib
from datetime import datetime
from urllib.parse import urljoin, urlparse
from dataclasses import dataclass, fields, asdict
from typing import Optional

import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DEFAULT_OUTPUT = "articles.xlsx"
REQUEST_TIMEOUT = 15          # seconds per request
DELAY_BETWEEN_REQUESTS = 2.0  # polite crawl delay in seconds
MAX_ARTICLES_PER_SITE = 100   # safety cap per site
MAX_RETRIES = 4               # max retries on rate-limit (429) responses
INITIAL_BACKOFF = 5.0         # initial wait (seconds) after first 429

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

# Minimal headers for the discovery session (homepage, RSS, sitemap).
# Deliberately simple — servers like WordPress return server-side-rendered HTML
# to simple clients, but may serve a JS-only shell when they see full Chrome
# headers (Accept: image/avif, Sec-Fetch-*, etc.).  These are what worked in
# early runs to reliably return 57+ article links from cardinalnews.org.
HEADERS_DISCOVERY = {
    "User-Agent": USER_AGENT,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# Full browser headers for the article session (used with cloudscraper).
# Cloudflare bot-detection on individual article pages checks for these.
HEADERS_BROWSER = {
    "User-Agent": USER_AGENT,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}

# Keep HEADERS as an alias pointing to the browser set (used in a few legacy spots)
HEADERS = HEADERS_BROWSER

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("news_scraper")

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Article:
    title: str
    author: Optional[str]
    date_published: Optional[str]
    body: str
    url: str
    source_site: str

    CSV_FIELDS = ["title", "author", "date_published", "body", "url", "source_site"]

# ---------------------------------------------------------------------------
# Page fetching
# ---------------------------------------------------------------------------

class PageFetcher:
    """Fetches page HTML with Cloudflare bypass support and optional Selenium fallback.

    Uses two sessions strategically:
    - discovery_session: plain requests.Session for homepages, RSS feeds, and
      sitemaps. These pages rarely trigger bot detection, and serving plain
      requests often returns better server-side-rendered HTML.
    - article_session: cloudscraper (if installed) for individual article pages,
      which are more likely to trigger per-page bot detection (429s). Falls back
      to requests if cloudscraper is not installed.

    Install cloudscraper for best results on Cloudflare-protected sites:
        pip install cloudscraper
    """

    def __init__(self, use_selenium: bool = False):
        self.driver = None
        # Plain session for discovery (homepage, RSS, sitemap) — minimal headers
        # so servers return server-side-rendered HTML rather than a JS-only shell
        self.discovery_session = requests.Session()
        self.discovery_session.headers.update(HEADERS_DISCOVERY)
        # Cloudscraper session for article fetches — full browser headers
        self.article_session = self._build_article_session()

        if use_selenium:
            self._init_selenium()

    def _build_article_session(self):
        """Create the best available HTTP session for article fetches."""
        try:
            import cloudscraper
            scraper = cloudscraper.create_scraper(
                browser={"browser": "chrome", "platform": "windows", "mobile": False}
            )
            scraper.headers.update(HEADERS_BROWSER)
            log.info("Using cloudscraper for article fetches (Cloudflare bypass enabled)")
            return scraper
        except ImportError:
            log.info(
                "cloudscraper not installed — using requests for article fetches. "
                "For better results on Cloudflare-protected sites: "
                "pip install cloudscraper"
            )
            session = requests.Session()
            session.headers.update(HEADERS_BROWSER)
            return session

    @property
    def session(self):
        """Backwards-compatible alias — returns the article session."""
        return self.article_session

    def _init_selenium(self):
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.chrome.options import Options
            from webdriver_manager.chrome import ChromeDriverManager

            options = Options()
            options.add_argument("--headless=new")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(f"user-agent={USER_AGENT}")
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            log.info("Selenium driver initialized (JS-rendering enabled)")
        except ImportError:
            log.warning(
                "Selenium not installed — falling back to requests only. "
                "Install with: pip install selenium webdriver-manager"
            )
        except Exception as e:
            log.warning(f"Could not start Selenium driver: {e}")

    def warm_up(self, site_url: str):
        """Visit the homepage to acquire cookies on both sessions.

        Many anti-bot systems expect a browser to have visited the root domain
        before navigating to individual pages.
        """
        try:
            log.info(f"  Warming up sessions on {site_url}...")
            resp = self.discovery_session.get(site_url, timeout=REQUEST_TIMEOUT)
            n_cookies = len(resp.cookies)
            # Share cookies with the article session too
            self.article_session.cookies.update(self.discovery_session.cookies)
            if resp.status_code == 200:
                log.info(f"  Sessions warmed up ({n_cookies} cookies set)")
            time.sleep(1.5)
        except Exception as e:
            log.warning(f"  Warm-up request failed: {e}")

    def fetch(
        self,
        url: str,
        referer: Optional[str] = None,
        use_discovery_session: bool = False,
        max_retries: Optional[int] = None,
    ) -> Optional[str]:
        """Return the HTML of a page, or None on failure.

        Automatically retries with exponential backoff on 429 (rate limit)
        responses and honours the Retry-After header when the server sends one.

        Args:
            url:                    The page to fetch.
            referer:                Optional Referer URL (simulates browser navigation).
            use_discovery_session:  If True, use the plain requests session instead
                                    of cloudscraper (better for homepages / RSS feeds
                                    that return richer HTML to simpler clients).
            max_retries:            Override the global MAX_RETRIES for this call.
                                    Pass 0 to try once with no retries (fail fast).
                                    Defaults to the global MAX_RETRIES setting.
        """
        retries = MAX_RETRIES if max_retries is None else max_retries

        # Set Referer header to mimic normal browser navigation
        headers = {}
        if referer:
            headers["Referer"] = referer
            headers["Sec-Fetch-Site"] = "same-origin"

        session = self.discovery_session if use_discovery_session else self.article_session

        # Try requests/cloudscraper first, with retry logic for 429s
        backoff = INITIAL_BACKOFF
        for attempt in range(1, retries + 2):  # +2: attempt 1 is always made
            try:
                resp = session.get(url, timeout=REQUEST_TIMEOUT, headers=headers)

                if resp.status_code == 429:
                    # Use Retry-After header if provided, otherwise backoff
                    retry_after = resp.headers.get("Retry-After")
                    if retry_after:
                        try:
                            wait = float(retry_after)
                        except ValueError:
                            wait = backoff
                    else:
                        wait = backoff

                    if attempt <= retries:
                        log.warning(
                            f"  Rate limited (429) on attempt {attempt}/{retries + 1}. "
                            f"Waiting {wait:.1f}s before retry..."
                        )
                        time.sleep(wait)
                        backoff = min(backoff * 2, 120)  # double, cap at 2 min
                        continue
                    else:
                        if retries > 0:
                            log.warning(
                                f"  Rate limited (429) on final attempt for {url} — skipping."
                            )
                        return None

                resp.raise_for_status()
                return resp.text

            except requests.RequestException as e:
                if "429" not in str(e):
                    log.warning(f"requests failed for {url}: {e}")
                    break  # non-429 error, don't retry — try Selenium
                if attempt <= retries:
                    log.warning(
                        f"  Rate limited (429) on attempt {attempt}/{retries + 1}. "
                        f"Waiting {backoff:.1f}s before retry..."
                    )
                    time.sleep(backoff)
                    backoff = min(backoff * 2, 120)
                else:
                    if retries > 0:
                        log.warning(f"  Rate limited after {retries + 1} attempts for {url}")
                    break

        # Fallback to Selenium if available
        if self.driver:
            try:
                self.driver.get(url)
                time.sleep(2)  # let JS render
                return self.driver.page_source
            except Exception as e:
                log.warning(f"Selenium failed for {url}: {e}")

        return None

    def close(self):
        if self.driver:
            self.driver.quit()

# ---------------------------------------------------------------------------
# RSS / Sitemap link discovery (preferred — machine-readable, no bot detection)
# ---------------------------------------------------------------------------

# Common RSS feed paths to probe
RSS_PATHS = ["/feed/", "/rss/", "/rss.xml", "/feed.xml", "/feeds/posts/default",
             "/atom.xml", "/news/feed/", "/feed/rss/"]

# Common sitemap paths to probe
SITEMAP_PATHS = ["/sitemap_index.xml", "/sitemap.xml", "/wp-sitemap.xml",
                 "/news-sitemap.xml", "/sitemap_news.xml", "/post-sitemap.xml"]


def _parse_rss_feed(xml_text: str, site_domain: str) -> tuple[list[str], list["Article"]]:
    """Parse an RSS/Atom feed XML, returning (links, articles).

    Articles are fully populated when the feed includes rich content
    (WordPress <content:encoded>, <dc:creator>, <pubDate>, etc.).
    Links are always returned so callers can fall back to per-page fetching
    when article content is sparse.
    """
    import json as _json

    soup = BeautifulSoup(xml_text, "lxml-xml")
    links: list[str] = []
    articles: list[Article] = []

    # ── RSS <item> entries ────────────────────────────────────────────────────
    for item in soup.find_all("item"):
        # URL
        link_tag = item.find("link")
        href = (link_tag.get_text(strip=True) if link_tag else "") or \
               (link_tag.get("href", "") if link_tag else "")
        if not href or not urlparse(href).netloc:
            continue
        article_url = href.split("?")[0].rstrip("/")
        links.append(article_url)

        # Title
        title_tag = item.find("title")
        title = title_tag.get_text(strip=True) if title_tag else None
        if not title:
            continue

        # Author (dc:creator namespace → BeautifulSoup sees it as "creator")
        creator_tag = item.find("creator")
        author = creator_tag.get_text(strip=True) if creator_tag else None

        # Date
        pubdate_tag = item.find("pubDate")
        date_str = pubdate_tag.get_text(strip=True) if pubdate_tag else None
        date_published = _normalize_date(date_str) if date_str else None

        # Body: prefer <content:encoded> (full HTML), fall back to <description>
        body = None
        encoded_tag = item.find("encoded")   # content:encoded
        if encoded_tag:
            raw_html = encoded_tag.get_text()
            content_soup = BeautifulSoup(raw_html, "lxml")
            body = _extract_body(content_soup)
            if not body:
                body = content_soup.get_text(separator="\n\n", strip=True)
        if not body:
            desc_tag = item.find("description")
            if desc_tag:
                raw = desc_tag.get_text(strip=True)
                desc_soup = BeautifulSoup(raw, "lxml")
                body = desc_soup.get_text(separator="\n\n", strip=True)

        if not body or len(body.strip()) < 50:
            continue

        articles.append(Article(
            title=title,
            author=author,
            date_published=date_published,
            body=body.strip(),
            url=article_url,
            source_site=site_domain,
        ))

    # ── Atom <entry> elements ─────────────────────────────────────────────────
    for entry in soup.find_all("entry"):
        link_tag = entry.find("link", href=True)
        if not link_tag:
            continue
        href = link_tag["href"]
        if not href or not urlparse(href).netloc:
            continue
        article_url = href.split("?")[0].rstrip("/")
        links.append(article_url)

        title_tag = entry.find("title")
        title = title_tag.get_text(strip=True) if title_tag else None
        if not title:
            continue

        author_tag = entry.find("name")  # <author><name>
        author = author_tag.get_text(strip=True) if author_tag else None

        pub_tag = entry.find("published") or entry.find("updated")
        date_published = _normalize_date(pub_tag.get_text(strip=True)) if pub_tag else None

        body = None
        content_tag = entry.find("content")
        if content_tag:
            raw = content_tag.get_text(strip=True)
            body_soup = BeautifulSoup(raw, "lxml")
            body = _extract_body(body_soup) or body_soup.get_text(separator="\n\n", strip=True)
        if not body:
            summary_tag = entry.find("summary")
            if summary_tag:
                body = BeautifulSoup(summary_tag.get_text(), "lxml").get_text(
                    separator="\n\n", strip=True
                )

        if not body or len(body.strip()) < 50:
            continue

        articles.append(Article(
            title=title,
            author=author,
            date_published=date_published,
            body=body.strip(),
            url=article_url,
            source_site=site_domain,
        ))

    return list(dict.fromkeys(links)), articles


def discover_links_from_rss(fetcher: "PageFetcher", site_url: str) -> list[str]:
    """Try to find article links via an RSS/Atom feed.

    Uses fail-fast mode (no retries per path) so the probe moves on quickly
    when a path doesn't exist or is rate-limited, rather than waiting 35s+ per path.
    """
    base = site_url.rstrip("/")
    domain = urlparse(site_url).netloc
    for path in RSS_PATHS:
        url = base + path
        try:
            xml_text = fetcher.fetch(url, use_discovery_session=True, max_retries=0)
            if not xml_text:
                continue
            links, _ = _parse_rss_feed(xml_text, domain)
            if links:
                log.info(f"  RSS feed found at {path} — {len(links)} links")
                return links
        except Exception as e:
            log.debug(f"  RSS probe failed for {path}: {e}")
    return []


def scrape_articles_from_wp_api(
    fetcher: "PageFetcher", site_url: str, site_domain: str
) -> list["Article"]:
    """Extract articles via the WordPress REST API (/wp-json/wp/v2/posts).

    This is the most reliable way to bulk-fetch WordPress posts:
    - Supports up to 100 posts per page with ?per_page=100
    - Uses ?page=N for real server-side pagination (not affected by RSS caching)
    - Returns JSON with author name embedded via ?_embed
    - Respects X-WP-Total / X-WP-TotalPages response headers

    Returns an empty list if the site doesn't expose the WP REST API.
    """
    import json as _json

    base = site_url.rstrip("/")
    api_base = f"{base}/wp-json/wp/v2/posts"

    # Probe: minimal request — just check the API exists and get total count.
    # Retries with backoff if the site is temporarily rate-limiting.
    probe_url = f"{api_base}?per_page=1"
    header_resp = None
    probe_backoff = 15.0
    for probe_attempt in range(1, 4):  # up to 3 probe attempts
        try:
            header_resp = fetcher.discovery_session.get(probe_url, timeout=REQUEST_TIMEOUT)
        except Exception as e:
            log.info(f"  WP API probe failed — skipping ({e})")
            return []
        log.info(f"  WP API probe status: {header_resp.status_code} (attempt {probe_attempt}/3)")
        if header_resp.status_code == 429:
            wait = float(header_resp.headers.get("Retry-After", probe_backoff))
            log.info(f"  WP API probe 429 — waiting {wait:.0f}s before retry...")
            time.sleep(wait)
            probe_backoff = min(probe_backoff * 2, 120)
            header_resp = None
            continue
        break

    if header_resp is None or header_resp.status_code != 200:
        status = header_resp.status_code if header_resp is not None else "no response"
        log.info(f"  WP API not available (HTTP {status}) — skipping")
        return []

    try:
        probe_data = _json.loads(header_resp.text)
        if not isinstance(probe_data, list):
            log.info(f"  WP API returned unexpected format — skipping")
            return []
        total_posts = int(header_resp.headers.get("X-WP-Total", 0))
    except _json.JSONDecodeError as e:
        log.info(f"  WP API returned non-JSON — skipping ({e})")
        return []

    cap = min(MAX_ARTICLES_PER_SITE, total_posts) if total_posts else MAX_ARTICLES_PER_SITE
    log.info(
        f"  WordPress REST API found at /wp-json/wp/v2/posts "
        f"— {total_posts} total posts, fetching up to {cap}"
    )

    # Give the server a moment after the probe before bulk requests.
    # Without this, probe + first page fire in the same second and the
    # larger page response immediately triggers rate limiting.
    time.sleep(3.0)

    # Auto-detect the largest batch size the server will accept without 429ing.
    # Try 50 first (fewer total requests), fall back to 10 if the site is strict.
    WP_PAGE_SIZE = 50
    _size_probe = fetcher.discovery_session.get(
        f"{api_base}?per_page={WP_PAGE_SIZE}&page=1", timeout=REQUEST_TIMEOUT
    )
    if _size_probe.status_code == 429:
        wait = float(_size_probe.headers.get("Retry-After", 15))
        log.info(f"  per_page=50 rate-limited — waiting {wait:.0f}s then trying per_page=10...")
        time.sleep(wait)
        WP_PAGE_SIZE = 10
        _size_probe = fetcher.discovery_session.get(
            f"{api_base}?per_page={WP_PAGE_SIZE}&page=1", timeout=REQUEST_TIMEOUT
        )

    all_articles: list[Article] = []
    author_id_map: dict[int, str] = {}
    raw_posts: list[dict] = []
    actual_total_pages = 1

    # Process the size-probe response as page 1 (avoid re-fetching it)
    if _size_probe.status_code == 200:
        try:
            first_page_posts = _json.loads(_size_probe.text)
            if isinstance(first_page_posts, list) and first_page_posts:
                actual_total_pages = int(
                    _size_probe.headers.get("X-WP-TotalPages", actual_total_pages)
                )
                raw_posts.extend(first_page_posts)
                log.info(
                    f"    API has {actual_total_pages} page(s) at {WP_PAGE_SIZE}/page"
                )
                log.info(
                    f"    API page 1/{actual_total_pages}: +{len(first_page_posts)} posts "
                    f"(total so far: {len(raw_posts)})"
                )
        except Exception as e:
            log.debug(f"  WP API page 1 parse failed: {e}")
    else:
        log.info(f"  WP API bulk fetch failed (HTTP {_size_probe.status_code}) — skipping")
        return []

    page = 2

    while len(raw_posts) < cap and page <= actual_total_pages:
        page_url = f"{api_base}?per_page={WP_PAGE_SIZE}&page={page}"

        # Fetch with exponential backoff on 429s
        resp = None
        backoff = 15.0
        for attempt in range(1, 5):  # up to 4 attempts
            try:
                resp = fetcher.discovery_session.get(page_url, timeout=REQUEST_TIMEOUT)
            except Exception as e:
                log.warning(f"    WP API page {page} request error: {e}")
                break
            if resp.status_code == 429:
                wait = float(resp.headers.get("Retry-After", backoff))
                log.warning(
                    f"    WP API 429 on page {page} (attempt {attempt}/4) "
                    f"— waiting {wait:.0f}s..."
                )
                time.sleep(wait)
                backoff = min(backoff * 2, 120)
                resp = None
                continue
            break  # success or non-429 error

        if resp is None or resp.status_code != 200:
            status = resp.status_code if resp is not None else "no response"
            log.info(f"    WP API page {page} failed (HTTP {status}) — stopping")
            break

        try:
            posts = _json.loads(resp.text)
            if not isinstance(posts, list) or not posts:
                break
        except Exception as e:
            log.debug(f"  WP API page {page} JSON parse failed: {e}")
            break

        raw_posts.extend(posts)
        log.info(
            f"    API page {page}/{actual_total_pages}: +{len(posts)} posts "
            f"(total so far: {len(raw_posts)})"
        )

        page += 1
        time.sleep(2.0)  # polite delay — reduce 429 risk on subsequent pages

    # Bulk-fetch author names for all unique author IDs in one request
    author_ids = list({p.get("author") for p in raw_posts if p.get("author")})
    if author_ids:
        try:
            ids_param = ",".join(str(i) for i in author_ids[:100])
            auth_resp = fetcher.discovery_session.get(
                f"{base}/wp-json/wp/v2/users?include={ids_param}&per_page=100",
                timeout=REQUEST_TIMEOUT,
            )
            if auth_resp.status_code == 200:
                for user in _json.loads(auth_resp.text):
                    author_id_map[user["id"]] = user.get("name", "")
        except Exception as e:
            log.debug(f"  Author lookup failed: {e}")

    # Build Article objects
    for post in raw_posts[:cap]:
        try:
            title = BeautifulSoup(
                post.get("title", {}).get("rendered", ""), "lxml"
            ).get_text(strip=True)
            if not title:
                continue

            article_url = post.get("link", "").split("?")[0].rstrip("/")
            if not article_url:
                continue

            date_published = _normalize_date(post.get("date", "")[:10])
            author = author_id_map.get(post.get("author", 0))

            content_html = post.get("content", {}).get("rendered", "")
            if content_html:
                content_soup = BeautifulSoup(content_html, "lxml")
                body = _extract_body(content_soup) or content_soup.get_text(
                    separator="\n\n", strip=True
                )
            else:
                excerpt_html = post.get("excerpt", {}).get("rendered", "")
                body = BeautifulSoup(excerpt_html, "lxml").get_text(
                    separator="\n\n", strip=True
                ) if excerpt_html else None

            if not body or len(body.strip()) < 50:
                continue

            all_articles.append(Article(
                title=title,
                author=author,
                date_published=date_published,
                body=body.strip(),
                url=article_url,
                source_site=site_domain,
            ))
        except Exception as e:
            log.debug(f"  WP API post parse error: {e}")
            continue

    log.info(
        f"  WordPress REST API complete — {len(all_articles)} articles "
        f"collected across {page} page(s)"
    )
    return all_articles


def fetch_post_by_wp_slug(
    fetcher: "PageFetcher",
    article_url: str,
    site_url: str,
    site_domain: str,
    author_cache: dict,
) -> Optional["Article"]:
    """Fetch a single WordPress article via the REST API using its URL slug.

    This is the key workaround for sites (like cardinalnews.org) that:
    - Block WP API pagination after page 1 due to aggressive rate limiting
    - Serve JavaScript-rendered pages where HTML title extraction fails

    A slug-based request (?slug=...) returns exactly 1 post — the same
    payload size as the probe that always succeeds — so it stays well under
    whatever rate limit is blocking bulk requests.
    """
    import json as _json

    slug = article_url.rstrip("/").split("/")[-1]
    if not slug or len(slug) < 3:
        return None

    base = site_url.rstrip("/")
    api_url = f"{base}/wp-json/wp/v2/posts?slug={slug}"

    try:
        resp = fetcher.discovery_session.get(api_url, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 429:
            # Rate limited — return immediately. The quota won't clear in 10s
            # and retrying just wastes time on every article.
            log.debug(f"  WP slug API 429 for {slug} — skipping")
            return None
        if resp.status_code != 200:
            return None

        posts = _json.loads(resp.text)
        if not isinstance(posts, list) or not posts:
            return None
        post = posts[0]

        # Title
        title = BeautifulSoup(
            post.get("title", {}).get("rendered", ""), "lxml"
        ).get_text(strip=True)
        if not title:
            return None

        # Date
        date_published = _normalize_date(post.get("date", "")[:10])

        # Author — use cache to avoid one request per article
        author_id = post.get("author", 0)
        author = author_cache.get(author_id)
        if author is None and author_id:
            try:
                auth_resp = fetcher.discovery_session.get(
                    f"{base}/wp-json/wp/v2/users/{author_id}",
                    timeout=REQUEST_TIMEOUT,
                )
                if auth_resp.status_code == 200:
                    user = _json.loads(auth_resp.text)
                    author = user.get("name", "")
                    author_cache[author_id] = author
            except Exception:
                author_cache[author_id] = ""  # cache the miss

        # Body
        content_html = post.get("content", {}).get("rendered", "")
        if content_html:
            content_soup = BeautifulSoup(content_html, "lxml")
            body = _extract_body(content_soup) or content_soup.get_text(
                separator="\n\n", strip=True
            )
        else:
            excerpt_html = post.get("excerpt", {}).get("rendered", "")
            body = BeautifulSoup(excerpt_html, "lxml").get_text(
                separator="\n\n", strip=True
            ) if excerpt_html else None

        if not body or len(body.strip()) < 50:
            return None

        return Article(
            title=title,
            author=author or None,
            date_published=date_published,
            body=body.strip(),
            url=article_url,
            source_site=site_domain,
        )

    except Exception as e:
        log.debug(f"  WP slug fetch failed for {slug}: {e}")
        return None


def scrape_articles_from_rss(
    fetcher: "PageFetcher", site_url: str, site_domain: str
) -> list["Article"]:
    """Extract full article data directly from an RSS/Atom feed, with pagination.

    WordPress RSS feeds include <content:encoded> (full article HTML),
    <dc:creator> (author), <pubDate>, and <title> — everything we need,
    without touching individual article pages.  This completely sidesteps
    per-page bot-detection and JS-rendered metadata issues.

    WordPress paginates RSS via ?paged=N (default 10 items/page).  We keep
    fetching pages until we hit MAX_ARTICLES_PER_SITE or get no new items.

    Returns an empty list if no feed is found or the feed lacks body content.
    """
    base = site_url.rstrip("/")
    for path in RSS_PATHS:
        feed_base_url = base + path
        try:
            # Page 1 — always fetch without ?paged= so it works on non-WordPress too
            xml_text = fetcher.fetch(feed_base_url, use_discovery_session=True, max_retries=0)
            if not xml_text:
                continue
            links, articles = _parse_rss_feed(xml_text, site_domain)
            if not links:
                continue  # not a valid feed

            if not articles:
                log.info(
                    f"  RSS feed at {path} found {len(links)} links "
                    f"but no extractable body content — will fetch pages individually"
                )
                return []

            log.info(f"  RSS feed found at {path} — paginating to collect all articles...")
            all_articles: list[Article] = list(articles)
            seen_urls: set[str] = {a.url for a in articles}

            # WordPress supports two feed pagination formats — try both:
            #   1. /feed/page/2/    (WP pretty permalinks — most reliable)
            #   2. /feed/?paged=2   (query-string — may be cached/ignored)
            # Probe page 2 with both formats to find which one works.
            base_no_slash = feed_base_url.rstrip("/")
            # Build candidate page-2 URLs
            format_a = f"{base_no_slash}/page/2/"               # pretty permalink
            sep = "&" if "?" in feed_base_url else "?"
            format_b = f"{feed_base_url}{sep}paged=2"           # query-string

            pagination_template = None  # will be set once we find the working format
            for candidate in (format_a, format_b):
                probe = fetcher.fetch(candidate, use_discovery_session=True, max_retries=0)
                if probe:
                    probe_links, _ = _parse_rss_feed(probe, site_domain)
                    new = [u for u in probe_links if u not in seen_urls]
                    if new:
                        pagination_template = candidate.replace("/2/", "/{page}/").replace(
                            "paged=2", "paged={page}"
                        )
                        log.info(f"  RSS pagination format: {pagination_template}")
                        # Process the articles from page 2 we already fetched
                        _, p2_articles = _parse_rss_feed(probe, site_domain)
                        new_arts = [a for a in p2_articles if a.url not in seen_urls]
                        for art in new_arts:
                            seen_urls.add(art.url)
                            all_articles.append(art)
                        log.info(f"    Page 2: +{len(new_arts)} articles (total: {len(all_articles)})")
                        break

            if pagination_template is None:
                log.info(
                    f"  RSS feed doesn't support pagination "
                    f"(neither /page/N/ nor ?paged=N returned new articles) — {len(all_articles)} articles total"
                )
            else:
                # Continue from page 3 onward
                page = 3
                consecutive_empty = 0
                while len(all_articles) < MAX_ARTICLES_PER_SITE:
                    page_url = pagination_template.format(page=page)
                    xml_text = fetcher.fetch(page_url, use_discovery_session=True, max_retries=0)
                    if not xml_text:
                        break

                    _, page_articles = _parse_rss_feed(xml_text, site_domain)
                    new_articles = [a for a in page_articles if a.url not in seen_urls]
                    if not new_articles:
                        consecutive_empty += 1
                        if consecutive_empty >= 2:
                            break
                        page += 1
                        continue

                    consecutive_empty = 0
                    for art in new_articles:
                        seen_urls.add(art.url)
                        all_articles.append(art)

                    log.info(
                        f"    Page {page}: +{len(new_articles)} articles "
                        f"(total so far: {len(all_articles)})"
                    )
                    time.sleep(0.5)
                    page += 1

            log.info(
                f"  RSS pagination complete — {len(all_articles)} total articles "
                f"(no individual page fetches needed)"
            )
            return all_articles

        except Exception as e:
            log.debug(f"  RSS article extraction failed for {path}: {e}")
    return []


def discover_links_from_sitemap(fetcher: "PageFetcher", site_url: str) -> list[str]:
    """Try to find article links via XML sitemap.

    Uses fail-fast mode for probing (no retries per path). Child sitemaps
    within a sitemap index get one retry since they're known to exist.
    """
    base = site_url.rstrip("/")
    domain = urlparse(site_url).netloc

    def _parse_sitemap(xml_text: str) -> list[str]:
        """Parse a sitemap XML; recursively follows sitemap index entries."""
        soup = BeautifulSoup(xml_text, "lxml-xml")
        links = []
        # Sitemap index — recurse into child sitemaps
        for loc in soup.find_all("sitemap"):
            child_loc = loc.find("loc")
            if child_loc:
                child_url = child_loc.get_text(strip=True)
                child_html = fetcher.fetch(child_url, use_discovery_session=True, max_retries=1)
                if child_html:
                    links.extend(_parse_sitemap(child_html))
                    time.sleep(0.5)
        # Regular sitemap — collect <url><loc> entries that look like articles
        for url_tag in soup.find_all("url"):
            loc = url_tag.find("loc")
            if loc:
                href = loc.get_text(strip=True)
                if href and urlparse(href).netloc == domain and looks_like_article_url(href):
                    links.append(href.split("?")[0].rstrip("/"))
        return links

    for path in SITEMAP_PATHS:
        url = base + path
        try:
            xml_text = fetcher.fetch(url, use_discovery_session=True, max_retries=0)
            if not xml_text or "<" not in xml_text:
                continue
            links = _parse_sitemap(xml_text)
            if links:
                log.info(f"  Sitemap found at {path} — {len(links)} article links")
                return list(dict.fromkeys(links))
        except Exception as e:
            log.debug(f"  Sitemap probe failed for {path}: {e}")
    return []

# ---------------------------------------------------------------------------
# Article link discovery (HTML-based fallback)
# ---------------------------------------------------------------------------

# Patterns that usually indicate an article URL on news sites
ARTICLE_PATH_PATTERNS = [
    r"/\d{4}/\d{2}/\d{2}/",        # /2024/01/15/...
    r"/\d{4}/\d{2}/",              # /2024/01/...
    r"/article[s]?/",              # /article/ or /articles/
    r"/story/",                    # /story/...
    r"/news/",                     # /news/...
    r"/local/",                    # /local/...
    r"/crime/",                    # /crime/...
    r"/politics/",                 # /politics/...
    r"/business/",                 # /business/...
    r"/sports/",                   # /sports/...
    r"/opinion/",                  # /opinion/...
    r"/community/",               # /community/...
    r"/entertainment/",           # /entertainment/...
    r"/education/",               # /education/...
    r"/health/",                  # /health/...
    r"/environment/",             # /environment/...
    r"/breaking/",                # /breaking/...
    r"/post/",                    # /post/...
]

# Paths to skip (not articles)
SKIP_PATTERNS = [
    r"/tag/", r"/category/", r"/author/", r"/page/",
    r"/search", r"/login", r"/signup", r"/subscribe",
    r"/contact", r"/about", r"/privacy", r"/terms",
    r"/advertise", r"/rss", r"/feed", r"/wp-admin",
    r"/wp-login", r"/cart", r"/account",
    r"\.(jpg|jpeg|png|gif|svg|pdf|mp4|mp3|zip)$",
]


def looks_like_article_url(url: str) -> bool:
    """Heuristic check: does this URL look like it points to a news article?"""
    path = urlparse(url).path.lower()

    # Skip obvious non-article paths
    for pattern in SKIP_PATTERNS:
        if re.search(pattern, path, re.IGNORECASE):
            return False

    # Check for article-like path patterns
    for pattern in ARTICLE_PATH_PATTERNS:
        if re.search(pattern, path, re.IGNORECASE):
            return True

    # Also accept long slugs (e.g., /some-headline-about-local-event)
    slug = path.rstrip("/").split("/")[-1] if path.rstrip("/") else ""
    if slug and "-" in slug and len(slug) > 20:
        return True

    return False


def discover_article_links(html: str, base_url: str) -> list[str]:
    """Find all probable article links on a page."""
    soup = BeautifulSoup(html, "lxml")
    base_domain = urlparse(base_url).netloc
    seen = set()
    links = []

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].strip()
        full_url = urljoin(base_url, href)

        # Only follow links on the same domain
        if urlparse(full_url).netloc != base_domain:
            continue

        # Normalize
        full_url = full_url.split("#")[0].split("?")[0].rstrip("/")

        if full_url in seen:
            continue
        seen.add(full_url)

        if looks_like_article_url(full_url):
            links.append(full_url)

    return links

# ---------------------------------------------------------------------------
# Article extraction
# ---------------------------------------------------------------------------

def extract_article(
    html: str,
    url: str,
    source_site: str,
    debug: bool = False,
) -> Optional[Article]:
    """Extract article data from a single article page.

    Args:
        html:        Raw HTML of the article page.
        url:         Canonical URL of the article.
        source_site: Domain of the source site.
        debug:       If True, log extra diagnostics when extraction fails.
    """
    soup = BeautifulSoup(html, "lxml")

    title = _extract_title(soup, html)
    if not title and len(html) > 5000:
        # lxml sometimes mangles <head> on non-standard HTML, dropping meta/title
        # tags entirely while keeping body content intact.  html.parser is more
        # lenient and often recovers the title when lxml fails.
        soup = BeautifulSoup(html, "html.parser")
        title = _extract_title(soup, html)

    # Absolute last resort: derive a readable title from the URL slug.
    # Headless CMS frontends (Next.js, etc.) often omit all standard title
    # elements from server-rendered HTML while still including the full
    # article body.  The URL slug is a reliable fallback.
    if not title and url:
        slug = url.rstrip("/").split("/")[-1]
        if slug and len(slug) > 10:
            title = slug.replace("-", " ").strip()
            title = " ".join(w.capitalize() for w in title.split())
            if debug:
                log.info(f"    [title] No standard title found — using URL slug: {title[:80]}")

    if not title:
        if debug:
            page_title = soup.find("title")
            page_title_text = page_title.get_text(strip=True) if page_title else "(no <title> tag)"
            body_len = len(soup.get_text(strip=True))
            log.info(
                f"    [debug] Extraction failed — no title found.\n"
                f"           <title> tag text : {page_title_text[:120]}\n"
                f"           Total page text  : {body_len} chars\n"
                f"           Looks like CF?   : {'yes' if _looks_like_cf_challenge(html) else 'no'}"
            )
        return None

    author = _extract_author(soup)
    date_published = _extract_date(soup)
    body = _extract_body(soup)

    if not body or len(body.strip()) < 100:
        if debug:
            body_len = len(body.strip()) if body else 0
            log.info(
                f"    [debug] Extraction failed — body too short.\n"
                f"           Title found      : {title[:80]}\n"
                f"           Body length      : {body_len} chars (need ≥100)\n"
                f"           Looks like CF?   : {'yes' if _looks_like_cf_challenge(html) else 'no'}"
            )
        return None  # Probably not a real article

    return Article(
        title=title.strip(),
        author=author.strip() if author else None,
        date_published=date_published.strip() if date_published else None,
        body=body.strip(),
        url=url,
        source_site=source_site,
    )


def _looks_like_cf_challenge(html: str) -> bool:
    """Heuristic: does this HTML look like a Cloudflare JS challenge page?"""
    lower = html.lower()
    indicators = [
        "checking your browser",
        "enable javascript",
        "cf-browser-verification",
        "jschl_vc",
        "jschl-answer",
        "cf-spinner",
        "just a moment",
        "ddos-guard",
        "ray id",
        "cloudflare",
    ]
    matches = sum(1 for ind in indicators if ind in lower)
    # Also check for very short pages (CF challenge pages are typically <5KB)
    if len(html) < 5000 and matches >= 1:
        return True
    return matches >= 2


def _extract_title(soup: BeautifulSoup, raw_html: str = "") -> Optional[str]:
    """Extract the article title using multiple strategies."""
    # Strategy 1: og:title meta tag (most reliable)
    og = soup.find("meta", attrs={"property": "og:title"})
    if og and og.get("content"):
        return og["content"]

    # Strategy 1b: twitter:title (common fallback on modern CMS)
    tw = soup.find("meta", attrs={"name": "twitter:title"})
    if tw and tw.get("content"):
        return tw["content"]

    # Strategy 2: <h1> tag (usually the headline)
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        return h1.get_text(strip=True)

    # Strategy 3: <title> tag
    title_tag = soup.find("title")
    if title_tag and title_tag.get_text(strip=True):
        text = title_tag.get_text(strip=True)
        # Remove common suffixes like " - Site Name" or " | Site Name"
        text = re.split(r"\s*[|\-–—]\s*", text)[0]
        return text

    # Strategy 4: JSON-LD "headline" field (reliable on well-structured sites)
    import json as _json
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = _json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            headline = data.get("headline") or data.get("name")
            if headline and isinstance(headline, str) and len(headline) > 5:
                return headline
        except Exception:
            continue

    # Strategy 5: Raw regex fallbacks on the original HTML string.
    # Catches titles that parsers drop when the HTML is malformed or when
    # the <head> is missing from the parsed DOM.
    if raw_html:
        # og:title in raw HTML
        m = re.search(r'property=["\']og:title["\'][^>]*content=["\']([^"\']+)["\']', raw_html, re.IGNORECASE)
        if not m:
            m = re.search(r'content=["\']([^"\']+)["\'][^>]*property=["\']og:title["\']', raw_html, re.IGNORECASE)
        if m:
            return m.group(1).strip()
        # <title> tag
        m = re.search(r'<title[^>]*>([^<]{5,})</title>', raw_html, re.IGNORECASE)
        if m:
            text = m.group(1).strip()
            return re.split(r"\s*[|\-–—]\s*", text)[0]
        # JSON-LD headline
        m = re.search(r'"headline"\s*:\s*"([^"]{5,})"', raw_html)
        if m:
            return m.group(1).strip()

    # Strategy 6: Broaden heading search — try <h2> and elements with
    # title-related CSS classes (headless CMS frontends often skip <h1>).
    h2 = soup.find("h2")
    if h2:
        text = h2.get_text(strip=True)
        if text and 10 < len(text) < 300:
            return text

    title_class_selectors = [
        '[class*="entry-title"]', '[class*="post-title"]',
        '[class*="article-title"]', '[class*="article__title"]',
        '[class*="article-header__title"]', '[class*="story-title"]',
        '[class*="headline"]', '[class*="page-title"]',
    ]
    for sel in title_class_selectors:
        el = soup.select_one(sel)
        if el:
            text = el.get_text(strip=True)
            if text and 10 < len(text) < 300:
                return text

    return None


def _extract_author(soup: BeautifulSoup) -> Optional[str]:
    """Extract the article author."""
    # Strategy 1: meta tags
    for attr_name in ["author", "article:author"]:
        meta = soup.find("meta", attrs={"name": attr_name}) or soup.find(
            "meta", attrs={"property": attr_name}
        )
        if meta and meta.get("content"):
            return meta["content"]

    # Strategy 2: JSON-LD structured data
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            import json
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            author = data.get("author")
            if isinstance(author, dict):
                return author.get("name")
            elif isinstance(author, list) and author:
                return author[0].get("name") if isinstance(author[0], dict) else str(author[0])
            elif isinstance(author, str):
                return author
        except (json.JSONDecodeError, AttributeError, IndexError):
            continue

    # Strategy 3: Common CSS class/attribute patterns
    author_selectors = [
        '[class*="author"]', '[class*="byline"]', '[rel="author"]',
        '[itemprop="author"]', '.writer', '.contributor',
    ]
    for selector in author_selectors:
        el = soup.select_one(selector)
        if el:
            text = el.get_text(strip=True)
            # Clean up "By John Smith" patterns
            text = re.sub(r"^[Bb]y\s+", "", text)
            if text and len(text) < 100:  # sanity check
                return text

    return None


def _extract_date(soup: BeautifulSoup) -> Optional[str]:
    """Extract the publication date."""
    # Strategy 1: meta tags
    date_meta_names = [
        "article:published_time", "og:article:published_time",
        "date", "pubdate", "publish-date", "DC.date.issued",
        "datePublished", "article:published",
    ]
    for name in date_meta_names:
        meta = soup.find("meta", attrs={"property": name}) or soup.find(
            "meta", attrs={"name": name}
        )
        if meta and meta.get("content"):
            return _normalize_date(meta["content"])

    # Strategy 2: <time> element
    time_el = soup.find("time", attrs={"datetime": True})
    if time_el:
        return _normalize_date(time_el["datetime"])
    time_el = soup.find("time")
    if time_el and time_el.get_text(strip=True):
        return time_el.get_text(strip=True)

    # Strategy 3: JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            import json
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            for key in ("datePublished", "dateCreated"):
                if key in data:
                    return _normalize_date(data[key])
        except (json.JSONDecodeError, AttributeError):
            continue

    return None


def _normalize_date(date_str: str) -> str:
    """Try to normalize a date string to YYYY-MM-DD format."""
    if not date_str:
        return date_str
    date_str = date_str.strip()

    # Already ISO format — grab just the date portion
    match = re.match(r"(\d{4}-\d{2}-\d{2})", date_str)
    if match:
        return match.group(1)

    # RFC 2822 format used in RSS <pubDate>: "Tue, 18 Mar 2026 12:00:00 +0000"
    # email.utils.parsedate handles the full range of RFC 2822 variants reliably
    try:
        from email.utils import parsedate
        parsed = parsedate(date_str)
        if parsed:
            return datetime(*parsed[:3]).strftime("%Y-%m-%d")
    except Exception:
        pass

    # Try other common formats
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %b %Y",
                "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return date_str


def _extract_body(soup: BeautifulSoup) -> Optional[str]:
    """Extract the main article body text."""
    # Remove script, style, nav, footer, aside, header elements
    for tag in soup.find_all(["script", "style", "nav", "footer", "aside",
                               "header", "form", "iframe", "noscript"]):
        tag.decompose()

    # Strategy 1: Look for article tag or common article containers
    article_selectors = [
        "article",
        '[itemprop="articleBody"]',
        '[class*="article-body"]',
        '[class*="article-content"]',
        '[class*="story-body"]',
        '[class*="story-content"]',
        '[class*="entry-content"]',
        '[class*="post-content"]',
        '[class*="post-body"]',
        '[class*="content-body"]',
        '[class*="article__body"]',
        '[class*="article__content"]',
        ".story-text",
        "#article-body",
        "#story-body",
    ]

    for selector in article_selectors:
        container = soup.select_one(selector)
        if container:
            paragraphs = container.find_all("p")
            if paragraphs:
                text = "\n\n".join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))
                if len(text) > 100:
                    return text

    # Strategy 2: Largest cluster of <p> tags
    # Find the parent element that contains the most paragraph text
    best_parent = None
    best_score = 0

    for p in soup.find_all("p"):
        parent = p.parent
        if parent:
            score = sum(
                len(child.get_text(strip=True))
                for child in parent.find_all("p", recursive=False)
            )
            if score > best_score:
                best_score = score
                best_parent = parent

    if best_parent and best_score > 100:
        paragraphs = best_parent.find_all("p")
        text = "\n\n".join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))
        return text

    return None

# ---------------------------------------------------------------------------
# Keyword extraction
# ---------------------------------------------------------------------------

_STOP_WORDS = frozenset("""
a about above after again against ago all also although always am an and
any are aren't as at be because been before being below between both but by
can can't cannot could couldn't did didn't do does doesn't doing don't down
during each even ever every few for from further get got had hadn't has
hasn't have haven't having he he'd he'll he's her here here's hers herself
him himself his how how's i i'd i'll i'm i've if in into is isn't it it's
its itself just know let like ll ll's made make many may me more most must
my myself new no nor not now of off on once only or other our ours ourselves
out over own re re's said same say says she she'd she'll she's should
shouldn't since so some still such than that that's the their theirs them
themselves then there there's these they they'd they'll they're they've this
those through to too under until up us ve very was wasn't we we'd we'll
we're we've were weren't what what's when when's where where's which while
who who's whom why why's will with won't would wouldn't year years you
you'd you'll you're you've your yours yourself yourselves been being having
also just still even like back only well already always often never later
early another either neither every across along among around upon between
another whether however therefore although though unless since while
until after before once already whether again also just now then there here
today tomorrow yesterday week weeks month months day days time times news
report reported reporting reports article articles story stories said says
told according local new first last one two three four five more most many
such part next each own old high long little good great small large big
""".split())


def _extract_keywords(articles: list[Article]) -> list:
    """
    Extract meaningful keywords from article titles and body text.

    Returns (word, count) tuples sorted by count descending,
    only including words that appear at least twice and are not stop words.
    """
    from collections import Counter
    counts: Counter = Counter()
    for article in articles:
        text = f"{article.title or ''} {article.body or ''}"
        for word in re.findall(r"[a-zA-Z]{4,}", text):
            lw = word.lower()
            if lw not in _STOP_WORDS:
                counts[lw] += 1
    return [(word, count) for word, count in counts.most_common() if count >= 2]


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_COL_WIDTHS = {
    "title": 52, "author": 22, "date_published": 15,
    "body": 90,  "url": 55,   "source_site": 28,
}
_HEADER_FILL = "D9E1F2"


def _style_header_row(ws, num_cols: int):
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1][:num_cols]:
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_keywords_sheet(ws, articles: list[Article]):
    from openpyxl.styles import Font, Alignment
    ws.append(["Keyword", "Count"])
    _style_header_row(ws, 2)
    for word, count in _extract_keywords(articles):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=word).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=2, value=count).font = Font(name="Arial", size=10)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A2"


def write_xlsx(articles: list[Article], output_path: str):
    """Write articles to an Excel file with Articles + Keywords sheets (overwrites)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fields = Article.CSV_FIELDS

    ws = wb.active
    ws.title = "Articles"
    ws.append(fields)
    _style_header_row(ws, len(fields))

    for article in articles:
        d = asdict(article)
        ws.append([d.get(f) or "" for f in fields])

    for col_idx, field in enumerate(fields, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(field, 20)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    ws_kw = wb.create_sheet("Keywords")
    _write_keywords_sheet(ws_kw, articles)

    wb.save(output_path)
    log.info(f"Wrote {len(articles)} articles + keywords to {output_path}")


def _append_to_xlsx(articles: list[Article], output_path: str):
    """
    Append new articles to an existing xlsx, skip URL duplicates,
    then rebuild the Keywords sheet from the full updated article list.
    """
    if not os.path.exists(output_path):
        write_xlsx(articles, output_path)
        return

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)

    if "Articles" not in wb.sheetnames:
        write_xlsx(articles, output_path)
        return

    ws = wb["Articles"]
    fields = Article.CSV_FIELDS
    url_col_idx = fields.index("url")  # 0-based

    existing_urls: set = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[url_col_idx]
        if val:
            existing_urls.add(str(val))

    new_articles = [a for a in articles if a.url not in existing_urls]
    if not new_articles:
        log.info(f"No new articles to append (all {len(articles)} already in {output_path})")
        return

    for article in new_articles:
        d = asdict(article)
        ws.append([d.get(f) or "" for f in fields])

    last_row = ws.max_row
    first_new = last_row - len(new_articles) + 1
    for col_idx in range(1, len(fields) + 1):
        for row in ws.iter_rows(min_row=first_new, max_row=last_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    # Rebuild Keywords from all rows now in the sheet
    all_articles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(fields, row))
        all_articles.append(Article(
            title=str(d.get("title") or ""),
            author=d.get("author"),
            date_published=d.get("date_published"),
            body=str(d.get("body") or ""),
            url=str(d.get("url") or ""),
            source_site=str(d.get("source_site") or ""),
        ))

    if "Keywords" in wb.sheetnames:
        del wb["Keywords"]
    ws_kw = wb.create_sheet("Keywords")
    _write_keywords_sheet(ws_kw, all_articles)

    wb.save(output_path)
    log.info(f"Appended {len(new_articles)} articles; Keywords sheet rebuilt in {output_path}")

# ---------------------------------------------------------------------------
# Main scraping orchestrator
# ---------------------------------------------------------------------------

def scrape_site(
    fetcher: PageFetcher,
    site_url: str,
    debug_html_path: Optional[str] = None,
    debug_article_path: Optional[str] = None,
    target_dates: Optional[set] = None,
) -> list[Article]:
    """Scrape all discoverable articles from a single news site."""
    log.info(f"Scraping: {site_url}")
    site_domain = urlparse(site_url).netloc

    # ── Step 0a: Try WordPress REST API (BEFORE warm_up) ─────────────────────
    # The warm_up homepage request counts against the same rate-limit budget as
    # the API.  If warm_up fires first, the immediately-following API probe gets
    # 429'd even with a delay.  Trying the API cold — before any homepage hit —
    # gives it the best chance of succeeding.
    log.info(f"  Trying WordPress REST API...")
    bulk_articles = scrape_articles_from_wp_api(fetcher, site_url, site_domain)

    # Warm up sessions now (acquires cookies for per-page fetching fallbacks)
    fetcher.warm_up(site_url)

    # ── Step 0b: Fall back to RSS feed (with pagination) ─────────────────────
    # Works for WordPress and non-WordPress sites. Pagination via ?paged=N may
    # be unreliable on cached setups, but RSS content is complete per item.
    if not bulk_articles:
        log.info(f"  Trying full-article RSS extraction...")
        bulk_articles = scrape_articles_from_rss(fetcher, site_url, site_domain)

    # Apply date filter if requested (--today / --date)
    if target_dates and bulk_articles:
        before = len(bulk_articles)
        bulk_articles = [
            a for a in bulk_articles
            if a.date_published and a.date_published[:10] in target_dates
        ]
        label = ", ".join(sorted(target_dates))
        log.info(f"  Date filter ({label}): kept {len(bulk_articles)}/{before} articles")
        if not bulk_articles:
            log.info(f"  No articles found for {label} via bulk methods.")

    # Deduplicate whatever bulk methods returned
    bulk_deduped: list[Article] = []
    seen_bulk_titles: set[str] = set()
    for art in bulk_articles[:MAX_ARTICLES_PER_SITE]:
        h = hashlib.md5(art.title.lower().encode()).hexdigest()
        if h not in seen_bulk_titles:
            seen_bulk_titles.add(h)
            bulk_deduped.append(art)
            log.info(f"    ✓ {art.title[:80]}")

    if bulk_deduped:
        log.info(f"  Bulk methods collected {len(bulk_deduped)} articles")

    # In date-filter mode, the RSS feed already covers the day's articles fully
    # (they're always in the most-recent 10).  Return immediately — no need to
    # fight rate limiters for historical bulk pagination.
    if target_dates:
        return bulk_deduped

    # If bulk methods gave us everything we need, return now
    if len(bulk_deduped) >= MAX_ARTICLES_PER_SITE:
        return bulk_deduped

    # Bulk methods either gave nothing, or only partial results (e.g. WP API
    # was rate-limited after page 1).  Either way, supplement by discovering
    # all article URLs and fetching the ones we don't already have.
    already_fetched_urls: set[str] = {a.url for a in bulk_deduped}
    if bulk_deduped:
        remaining_needed = MAX_ARTICLES_PER_SITE - len(bulk_deduped)
        log.info(
            f"  Bulk methods got {len(bulk_deduped)}/{MAX_ARTICLES_PER_SITE} articles — "
            f"discovering remaining {remaining_needed} via URL discovery..."
        )

    # ── Link discovery: try in order of reliability ──────────────────────────
    # 1. RSS feed (designed for machines, no bot detection, most reliable)
    log.info(f"  Trying RSS feed discovery...")
    article_urls = discover_links_from_rss(fetcher, site_url)

    # 2. Sitemap XML (also machine-readable, good coverage of all articles)
    if not article_urls:
        log.info(f"  Trying sitemap discovery...")
        article_urls = discover_links_from_sitemap(fetcher, site_url)

    # 3. HTML scraping — uses MINIMAL headers (discovery session) so the server
    #    returns server-side-rendered HTML rather than a JS-only shell
    if not article_urls:
        log.info(f"  Falling back to HTML link scraping...")
        html = fetcher.fetch(site_url, use_discovery_session=True)
        if not html:
            log.error(f"Could not fetch {site_url}")
            return []

        # Debug: save raw HTML so user can inspect what the server returned
        if debug_html_path:
            with open(debug_html_path, "w", encoding="utf-8") as f:
                f.write(html)
            log.info(f"  Raw homepage HTML saved to: {debug_html_path}")
            log.info(f"  (Open this file to check if a Cloudflare challenge page was returned)")

        article_urls = discover_article_links(html, site_url)
        log.info(f"  Found {len(article_urls)} candidate article links via HTML")

        if not article_urls:
            # Try common section pages
            section_paths = ["/news", "/local", "/local-news", "/latest", "/stories"]
            for path in section_paths:
                section_url = urljoin(site_url, path)
                section_html = fetcher.fetch(section_url, referer=site_url,
                                             use_discovery_session=True)
                if section_html:
                    extra = discover_article_links(section_html, site_url)
                    article_urls.extend(extra)
                    time.sleep(DELAY_BETWEEN_REQUESTS)
            article_urls = list(dict.fromkeys(article_urls))
            log.info(f"  After checking sections: {len(article_urls)} candidate links")

    # Remove URLs already covered by bulk methods (WP API / RSS)
    article_urls = [u for u in article_urls if u not in already_fetched_urls]
    log.info(
        f"  Total article links discovered: {len(article_urls)} "
        f"(excluding {len(already_fetched_urls)} already collected)"
    )

    # Cap: only fetch as many as we still need
    remaining_slots = MAX_ARTICLES_PER_SITE - len(bulk_deduped)
    article_urls = article_urls[:remaining_slots]

    # Detect whether the WP REST API is available for this site.
    # Detect whether the WP REST API is available for this site.
    # We know it's available if the bulk WP API call returned anything,
    # or if we can confirm with a lightweight probe.
    wp_api_available = any(
        a.source_site == site_domain for a in bulk_deduped
    ) and bool(bulk_deduped)
    # Cheaper check: just try a quick probe if we're about to fetch individual pages
    if not wp_api_available and article_urls:
        try:
            _probe = fetcher.discovery_session.get(
                f"{site_url.rstrip('/')}/wp-json/wp/v2/posts?per_page=1",
                timeout=5,
            )
            wp_api_available = _probe.status_code == 200
            if wp_api_available:
                log.info(f"  WP API available — will fetch remaining articles by slug")
        except Exception:
            pass

    # Author ID → name cache shared across all slug fetches
    wp_author_cache: dict[int, str] = {}

    # Fetch and extract each article
    articles = list(bulk_deduped)          # start with what bulk methods gave us
    seen_titles = {                        # pre-seed dedup set from bulk articles
        hashlib.md5(a.title.lower().encode()).hexdigest() for a in bulk_deduped
    }
    current_delay = DELAY_BETWEEN_REQUESTS  # adaptive delay
    last_successful_url = site_url  # track last page visited for Referer chain
    debug_article_saved = False  # only save the first one

    for i, article_url in enumerate(article_urls, 1):
        log.info(f"  [{i}/{len(article_urls)}] Fetching: {article_url}")
        time.sleep(current_delay)

        article = None

        # ── Strategy A: WP REST API by slug ──────────────────────────────────
        # Lightweight (1 post per request), avoids rate limits that block bulk
        # pagination, and returns clean structured data with proper titles.
        if wp_api_available:
            article = fetch_post_by_wp_slug(
                fetcher, article_url, site_url, site_domain, wp_author_cache
            )
            if article:
                log.info(f"    ✓ (via WP API slug): {article.title[:80]}")

        # ── Strategy B: HTML fetch + extraction ───────────────────────────────
        if not article:
            article_html = fetcher.fetch(article_url, referer=last_successful_url)
            if not article_html:
                current_delay = min(current_delay * 1.5, 30)
                log.info(f"    Increased delay to {current_delay:.1f}s after failure")
                continue

            last_successful_url = article_url

            if debug_article_path and not debug_article_saved:
                try:
                    with open(debug_article_path, "w", encoding="utf-8") as f:
                        f.write(article_html)
                    log.info(f"    [debug] Raw article HTML saved to: {debug_article_path}")
                    debug_article_saved = True
                except Exception as e:
                    log.warning(f"    [debug] Could not save article HTML: {e}")

            article = extract_article(article_html, article_url, site_domain, debug=True)

        if article:
            # Deduplicate by title
            title_hash = hashlib.md5(article.title.lower().encode()).hexdigest()
            if title_hash not in seen_titles:
                seen_titles.add(title_hash)
                articles.append(article)
                log.info(f"    ✓ Extracted: {article.title[:80]}")
            else:
                log.info(f"    ⊘ Duplicate skipped: {article.title[:60]}")
        else:
            log.info(f"    ✗ Could not extract article content")

    if bulk_deduped and len(articles) > len(bulk_deduped):
        log.info(
            f"  Combined: {len(bulk_deduped)} from bulk + "
            f"{len(articles) - len(bulk_deduped)} from individual fetches "
            f"= {len(articles)} total"
        )
    return articles

# ---------------------------------------------------------------------------
# Google Sheets integration
# ---------------------------------------------------------------------------

def parse_sheet_id(sheet_input: str) -> str:
    """Extract a Google Sheets ID from a full URL or return the raw ID."""
    # Full URL:  https://docs.google.com/spreadsheets/d/SHEET_ID/edit#gid=0
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_input)
    if match:
        return match.group(1)
    # Already a bare ID (alphanumeric + dashes/underscores, typically 44 chars)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", sheet_input):
        return sheet_input
    raise ValueError(
        f"Could not parse a Google Sheet ID from: {sheet_input}\n"
        "Provide either a full Google Sheets URL or the sheet ID."
    )


def parse_gid_from_url(sheet_input: str) -> Optional[str]:
    """Extract the gid parameter from a Google Sheets URL, if present."""
    match = re.search(r"[#&?]gid=(\d+)", sheet_input)
    return match.group(1) if match else None


def fetch_urls_from_google_sheet(
    sheet_input: str,
    tab_name: Optional[str] = None,
    column: str = "A",
) -> list[str]:
    """
    Read URLs from a Google Sheet.

    The sheet must be shared so that "Anyone with the link" can view it
    (Share → General access → Anyone with the link → Viewer).

    Args:
        sheet_input: A full Google Sheets URL or just the sheet ID.
        tab_name:    Specific tab/sheet name to read (default: first tab).
                     Ignored if the URL already contains a gid.
        column:      Column letter containing the URLs (default: "A").

    Returns:
        A list of URL strings found in that column.
    """
    sheet_id = parse_sheet_id(sheet_input)

    # Build the CSV export URL
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

    # Determine which tab (gid) to use
    gid = parse_gid_from_url(sheet_input)
    if gid:
        export_url += f"&gid={gid}"
    elif tab_name:
        # Google Sheets doesn't support tab-name export directly, so we
        # first fetch the HTML page to discover the gid for the named tab.
        gid = _resolve_tab_gid(sheet_id, tab_name)
        if gid is not None:
            export_url += f"&gid={gid}"
        else:
            log.warning(
                f"Could not find tab '{tab_name}' — falling back to first tab."
            )

    log.info(f"Fetching Google Sheet (column {column}): {export_url}")

    resp = requests.get(export_url, timeout=REQUEST_TIMEOUT)
    if resp.status_code == 403:
        raise PermissionError(
            "Cannot access the Google Sheet. Make sure it is shared so that "
            "'Anyone with the link' can view it.\n"
            "  → Open the sheet → Share → General access → "
            "'Anyone with the link' → Viewer"
        )
    resp.raise_for_status()

    # Parse the CSV and pull the target column
    col_index = _column_letter_to_index(column)
    reader = csv.reader(io.StringIO(resp.text))

    urls: list[str] = []
    for row_num, row in enumerate(reader):
        if col_index >= len(row):
            continue
        cell = row[col_index].strip()
        if not cell:
            continue
        # Skip header-like rows
        if row_num == 0 and not cell.startswith("http") and "." not in cell:
            continue
        # Normalize
        if not cell.startswith("http"):
            cell = "https://" + cell
        # Basic URL validation
        parsed = urlparse(cell)
        if parsed.scheme in ("http", "https") and "." in parsed.netloc:
            urls.append(cell)

    log.info(f"  Found {len(urls)} URLs in the Google Sheet")
    return urls


def fetch_url_groups_from_google_sheet(
    sheet_input: str,
    url_column: str = "A",
    output_column: str = "B",
    tab_name: Optional[str] = None,
    default_output: str = DEFAULT_OUTPUT,
) -> dict:
    """
    Read URLs and their target output CSV filenames from a Google Sheet.

    Column layout (letters are configurable):
        URL column (default A)  → the news site URL
        Output column (default B) → the CSV filename for that site's articles
                                    (e.g. "virginia.csv", "national.csv")

    Rows with an empty output column fall back to *default_output*.

    Returns:
        An ordered dict of  { output_filename: [url1, url2, ...] }
        so callers can iterate and write one CSV per group.
    """
    sheet_id = parse_sheet_id(sheet_input)

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    gid = parse_gid_from_url(sheet_input)
    if gid:
        export_url += f"&gid={gid}"
    elif tab_name:
        gid = _resolve_tab_gid(sheet_id, tab_name)
        if gid is not None:
            export_url += f"&gid={gid}"
        else:
            log.warning(
                f"Could not find tab '{tab_name}' — falling back to first tab."
            )

    log.info(
        f"Fetching Google Sheet (url_col={url_column}, "
        f"output_col={output_column}): {export_url}"
    )

    resp = requests.get(export_url, timeout=REQUEST_TIMEOUT)
    if resp.status_code == 403:
        raise PermissionError(
            "Cannot access the Google Sheet. Make sure it is shared so that "
            "'Anyone with the link' can view it.\n"
            "  → Open the sheet → Share → General access → "
            "'Anyone with the link' → Viewer"
        )
    resp.raise_for_status()

    url_col_idx = _column_letter_to_index(url_column)
    out_col_idx = _column_letter_to_index(output_column)

    reader = csv.reader(io.StringIO(resp.text))

    # Use an ordered dict so groups appear in sheet order
    from collections import OrderedDict
    groups: dict = OrderedDict()

    for row_num, row in enumerate(reader):
        # --- URL cell ---
        if url_col_idx >= len(row):
            continue
        url_cell = row[url_col_idx].strip()
        if not url_cell:
            continue
        # Skip a header row (first row that looks like a label, not a URL)
        if row_num == 0 and not url_cell.startswith("http") and "." not in url_cell:
            continue
        if not url_cell.startswith("http"):
            url_cell = "https://" + url_cell
        parsed = urlparse(url_cell)
        if parsed.scheme not in ("http", "https") or "." not in parsed.netloc:
            continue  # not a valid URL

        # --- Output file cell ---
        if out_col_idx < len(row):
            out_cell = row[out_col_idx].strip()
        else:
            out_cell = ""

        if not out_cell:
            out_cell = default_output

        # Ensure it ends with .csv
        if not out_cell.lower().endswith(".csv"):
            out_cell += ".csv"

        groups.setdefault(out_cell, [])
        groups[out_cell].append(url_cell)

    total_urls = sum(len(v) for v in groups.values())
    log.info(
        f"  Found {total_urls} URLs across {len(groups)} output group(s): "
        + ", ".join(f"{k}({len(v)})" for k, v in groups.items())
    )
    return groups


def _resolve_tab_gid(sheet_id: str, tab_name: str) -> Optional[str]:
    """Try to discover the gid for a named tab by fetching the sheet HTML."""
    try:
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
        # Look for the tab name and its associated gid in the page source
        # Google embeds something like: {"name":"MyTab","gid":"123456"}
        pattern = re.escape(tab_name)
        match = re.search(
            rf'"name"\s*:\s*"{pattern}"[^}}]*"gid"\s*:\s*"(\d+)"', resp.text
        )
        if match:
            return match.group(1)
        # Also try the reverse order
        match = re.search(
            rf'"gid"\s*:\s*"(\d+)"[^}}]*"name"\s*:\s*"{pattern}"', resp.text
        )
        if match:
            return match.group(1)
    except Exception as e:
        log.warning(f"Could not resolve tab name '{tab_name}': {e}")
    return None


def _column_letter_to_index(letter: str) -> int:
    """Convert a column letter (A, B, ..., Z, AA, ...) to a 0-based index."""
    letter = letter.upper().strip()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Scrape articles from local news websites and export to CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s -u https://www.localnews.com
  %(prog)s -u https://site1.com https://site2.com -o my_articles.csv
  %(prog)s -f sites.txt
  %(prog)s --sheet "https://docs.google.com/spreadsheets/d/SHEET_ID/edit"
  %(prog)s --sheet SHEET_ID --sheet-tab "Local News" --sheet-column B
  %(prog)s --sheet SHEET_ID --sheet-output-column B   # col A=URL, col B=output csv name
  %(prog)s                        # interactive mode
        """,
    )
    parser.add_argument(
        "-u", "--urls", nargs="+", metavar="URL",
        help="One or more news site URLs to scrape",
    )
    parser.add_argument(
        "-f", "--file", metavar="FILE",
        help="Path to a text file with one URL per line",
    )
    parser.add_argument(
        "--sheet", metavar="SHEET_URL_OR_ID",
        help="Google Sheets URL or sheet ID containing site URLs "
             "(sheet must be shared as 'Anyone with the link')",
    )
    parser.add_argument(
        "--sheet-tab", metavar="TAB_NAME", default=None,
        help="Name of the tab/sheet to read (default: first tab)",
    )
    parser.add_argument(
        "--sheet-column", metavar="COLUMN", default="A",
        help="Column letter containing the URLs (default: A)",
    )
    parser.add_argument(
        "--sheet-output-column", metavar="COLUMN", default=None,
        help="Column letter in the Google Sheet that specifies the output CSV "
             "filename for each site (e.g. B). When provided, sites that share "
             "the same filename in that column are scraped together and saved to "
             "that CSV — so one run can produce multiple output files "
             "(e.g. virginia.csv, national.csv). Rows with an empty cell in this "
             "column fall back to the --output filename.",
    )
    parser.add_argument(
        "-o", "--output", default=DEFAULT_OUTPUT,
        help=f"Output Excel file path (default: {DEFAULT_OUTPUT}). "
             "The file will contain an Articles sheet and a Keywords sheet.",
    )
    parser.add_argument(
        "--selenium", action="store_true",
        help="Enable Selenium for JavaScript-rendered pages",
    )
    parser.add_argument(
        "--max-articles", type=int, default=MAX_ARTICLES_PER_SITE,
        help=f"Max articles to scrape per site (default: {MAX_ARTICLES_PER_SITE})",
    )
    parser.add_argument(
        "--delay", type=float, default=DELAY_BETWEEN_REQUESTS,
        help=f"Base delay between requests in seconds; auto-increases "
             f"on rate limits (default: {DELAY_BETWEEN_REQUESTS})",
    )
    parser.add_argument(
        "--retries", type=int, default=MAX_RETRIES,
        help=f"Max retries per request on 429 rate-limit errors "
             f"(default: {MAX_RETRIES})",
    )
    parser.add_argument(
        "--debug-html", metavar="FILE", default=None,
        help="Save the raw homepage HTML to FILE for inspection. "
             "Useful for diagnosing why link discovery finds 0 articles "
             "(e.g. to check if a Cloudflare challenge page is being returned). "
             "Only fires when HTML fallback discovery is used (i.e. RSS/sitemap "
             "discovery did not find any links).",
    )
    parser.add_argument(
        "--debug-article", metavar="FILE", default=None,
        help="Save the raw HTML of the first article fetch to FILE for inspection. "
             "Useful for diagnosing why articles are fetched successfully but "
             "content extraction still fails (e.g. Cloudflare JS challenge page "
             "returned instead of real article content). The file saves to the "
             "folder where you run the script.",
    )
    parser.add_argument(
        "--today", action="store_true",
        help="Only collect articles published over the last --days days. "
             "Ideal for daily scheduled runs — avoids bulk pagination entirely, "
             "fast, rate-limit-safe, and appends new rows to the CSV without "
             "duplicating articles from previous runs.",
    )
    parser.add_argument(
        "--days", type=int, default=2,
        help="Number of days to look back when using --today (default: 2 = today and "
             "yesterday). Catches any articles missed by the previous day's run.",
    )
    parser.add_argument(
        "--date", metavar="YYYY-MM-DD", default=None,
        help="Only collect articles published on a specific date (e.g. 2026-03-20). "
             "Same as --today but for any date you choose.",
    )

    # ── Dashboard / GitHub export ──────────────────────────────────────────
    parser.add_argument(
        "--export-json", metavar="PATH", default=None,
        help="After scraping, write a dashboard-ready articles.json to this path "
             "(e.g. ~/news-dashboard/data/articles.json). The JSON includes articles, "
             "keyword counts with universe matches, source totals, and a date histogram. "
             "Can be used alongside --output (xlsx) or on its own.",
    )
    parser.add_argument(
        "--git-push", action="store_true",
        help="After writing --export-json, run 'git add . && git commit && git push' "
             "in the repo directory so GitHub Pages picks up the new data automatically. "
             "Requires git to be installed and the repo already cloned locally.",
    )
    parser.add_argument(
        "--universes-sheet", metavar="SHEET_URL_OR_ID", default=None,
        help="Google Sheets URL or ID containing the off-the-shelf voter targeting "
             "universes to match against keywords. The sheet should have at least two "
             "columns: universe name and description (or keywords). Used when building "
             "--export-json output.",
    )
    return parser.parse_args()


def get_urls_interactive() -> list[str]:
    """Prompt user for URLs when none are provided via CLI."""
    print("\n╔══════════════════════════════════════════╗")
    print("║       Local News Article Scraper         ║")
    print("╚══════════════════════════════════════════╝\n")
    print("Enter news site URLs one per line.")
    print("Press Enter on an empty line when done.\n")

    urls = []
    while True:
        try:
            url = input("  URL: ").strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not url:
            break
        if not url.startswith("http"):
            url = "https://" + url
        urls.append(url)

    return urls


# ---------------------------------------------------------------------------
# Universe loading + keyword matching
# ---------------------------------------------------------------------------

# Full Poseidon OTS Inventory (211 universes, sourced from the Pantheon sheet).
# Used as the default when --universes-sheet is not provided.
_OTS_UNIVERSES_RAW = [
    "General Turnout", "Primary Turnout", "Likely ABEV Voter",
    "Likely & Has ABEV Voter", "Urbanicity", "Voter Propensity",
    "Likely Donors", "Likely Activists", "Political Party",
    "Military Relationship", "Veteran", "Pro2A", "Small Biz Owners",
    "Parents", "Kids in Household", "Home Owners", "Home Renters",
    "Streaming Only", "TV Viewer", "Word of Month", "Education", "NIMBY",
    "Social Media", "Household Language", "Family Generation", "Generation",
    "Prolife", "Walkable Households", "Mailable Households", "Cleopatra",
    "Reducing crime", "Cost of Living", "Health care costs",
    "Outdoorsman/Conservationists", "Inconsistent Voters", "Trump not midterms",
    "Economically Stressed", "Independent Women", "Independent Men",
    "ESL Speakers", "Low News Intake", "High News Intake", "Cablezones",
    "Likely Survey Taker", "Ideology", "Influencers", "Movers", "Party Switcher",
    "Affordable Housing Priority", "Anti-CRT / Parental Rights in Education",
    "Anti-Regulation / Deregulation", "Anti-Tax / Tax Cut Supporters",
    "Anti-Vaccine / Vaccine Skeptic", "Anti-War / Non-Interventionist",
    "Back the Blue / Pro-Police", "Border Community Voters",
    "Broadband / High-Speed Internet", "Caregiver Households",
    "Casino/Gambling/Sports Betting", "Charitable Donors (Non-Political)",
    "Chronic Condition Households", "Cleopatra (Composite Persuasion Model)",
    "Clergy / Religious Workers", "Climate / Environment Priority",
    "College Town Voters", "College-Age Voters (18-22)", "College-Educated Men",
    "College-Educated Women", "Community Leaders / Local Influencers",
    "Concealed Carry Permit Holders", "Construction / Trades Workers",
    "Cost of Living Sensitive", "Country Music Listeners",
    "Craft Beer / Wine Enthusiasts", "Criminal Justice Reform",
    "Cross-Primary Voters", "CTV / OTT Targetable", "Disability Status",
    "Divorced Voters", "Drop-Off Voters", "Drug Policy Reform / Legalization",
    "Early / Absentee Voter Propensity", "Economic Growth / Jobs Priority",
    "Economically Stressed", "Education Priority",
    "Election Integrity / Voter ID Supporters",
    "Electric / Hybrid Vehicle Owners", "Electric Utility Rate Payers",
    "Empty Nesters", "Energy Assistance Recipients (LIHEAP)",
    "Energy Sector Workers", "Evangelical Christian", "Executive / C-Suite",
    "Farmers / Agricultural Workers", "Female Veterans",
    "Financial Services Workers", "First Responders (Police, Fire, EMS)",
    "First-Time Voters", "Flood Zone / Disaster-Prone Area",
    "Free Speech / Anti-Censorship", "Frequent Travelers", "Gamers (PC/Console)",
    "Gardeners", "Generation (Gen Z, Millennial, Gen X, Boomer, Silent)",
    "Gun Control Supporters", "Gun Owners (Consumer Data)", "HBCU Alumni",
    "Health Care Costs Priority", "Health Insurance Status (Insured/Uninsured)",
    "Healthcare Workers", "High News Intake", "Home Owners", "Home Renters",
    "Homeschool Parents", "Ideology Score (Liberal-Conservative)",
    "Immigration Hawks", "Immigration Reform / Pro-Immigrant",
    "Infrastructure Priority", "Kids in Household", "Latino Men",
    "Law Enforcement Family", "Legal Professionals", "Likely ABEV Voter",
    "Likely Activists", "Likely Donors (Political)", "Luxury Voters",
    "Manufacturing Workers", "Marijuana Legalization Supporters",
    "Medicaid Expansion Supporters", "Medicaid Recipients",
    "Medicare-Eligible (65+)", "Midterm Turnout Score", "Military (Active Duty)",
    "Military Base Adjacent", "Military Relationship (Household)",
    "Military Spouse", "Minimum Wage Increase Supporters",
    "Multi-Generational Households", "NASCAR / Motorsports Fans",
    "Near-Retirees (55-64)", "Neighborhood Association Active",
    "Neighborhood Watch / Community Safety", "Newly Registered Voters",
    "NIMBY (Not In My Back Yard)", "No Internet / Digital Divide",
    "Non-College Men", "Non-College Women", "Non-Profit Workers",
    "NPR Listeners / Donors", "Off-Year / Odd-Year Election Turnout",
    "Online Privacy Concerned", "Outdoorsman / Conservationists",
    "Parents (Any Children)", "Parents of School-Age Children (6-17)",
    "Parents of Young Children (0-5)", "Party Switcher", "Pet Owners",
    "Populist Score", "Post-9/11 Veterans",
    "Private / Parochial School Parents", "Pro-Choice / Abortion Rights",
    "Pro-Defense / Strong Military", "Pro-Life / Anti-Abortion",
    "Pro-Second Amendment / Gun Rights", "Pro-Union / Labor Rights",
    "PTA / School Board Engaged", "Public School Parents",
    "Real Estate Professionals", "Reducing Crime Priority",
    "Restaurant / Hospitality Workers", "Retail Workers", "Retired Voters",
    "Right-to-Work Supporters", "Rural Women", "Sandwich Generation",
    "School Choice Supporters", "Single Parents", "Small Business Owners",
    "Small Town Voters", "Smart Home / IoT Adopters",
    "SNAP / Food Assistance Recipients", "Social Media Heavy Users",
    "Social Media Non-Users", "Social Security / Medicare Defenders",
    "Social Security Recipients", "Special Election Turnout",
    "Straight-Ticket Voters", "Streaming Only (Cord-Cutters)",
    "Student Loan Holders", "Substance Abuse Impact", "Suburban Women",
    "Super Voters", "Teachers / Educators", "Tech Industry Workers",
    "Text Message Responsive", "Ticket Splitters",
    "Tough on Drugs / Anti-Legalization", "Transportation / Trucking",
    "Trump Support Score", "TV Viewer (Broadcast/Cable)",
    "Union Members (Current)", "Urbanicity (Urban/Suburban/Rural)",
    "VA Healthcare Users", "Veterans", "Video Game Players",
    "Vietnam-Era Veterans", "Walkability", "White Working-Class Voters",
    "Women of Color", "YIMBY (Yes In My Back Yard)",
]


def _build_universe_index(names: list) -> list:
    """Convert a list of universe name strings into matchable dicts."""
    result = []
    for name in names:
        if not name:
            continue
        norm = re.sub(r"[^a-z0-9 ]", " ", name.lower())
        tokens = frozenset(t for t in norm.split() if len(t) >= 3)
        result.append({"name": name, "norm": norm, "tokens": tokens})
    return result


# Pre-built index from the hardcoded OTS list
_DEFAULT_UNIVERSES = _build_universe_index(_OTS_UNIVERSES_RAW)


def _load_universes(sheet_input: Optional[str]) -> list:
    """
    Load voter-targeting universes.

    If *sheet_input* is None, returns the built-in Poseidon OTS list (211
    universes).  Otherwise tries to fetch from the given Google Sheet URL/ID
    (first column = universe name; second column = optional description).
    Falls back to the built-in list on any fetch error.
    """
    if not sheet_input:
        return _DEFAULT_UNIVERSES

    sheet_id = parse_sheet_id(sheet_input)
    export_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    )
    try:
        resp = requests.get(export_url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except Exception as e:
        log.warning(f"Could not load universes sheet ({e}); using built-in OTS list.")
        return _DEFAULT_UNIVERSES

    reader = csv.reader(io.StringIO(resp.text))
    names = []
    for row_num, row in enumerate(reader):
        if row_num < 2:       # skip "OTS Inventory Catalog" + "Model Name" header
            continue
        name = row[0].strip() if row else ""
        desc = row[1].strip() if len(row) > 1 else ""
        if name:
            names.append(f"{name} {desc}".strip() if desc else name)

    if not names:
        log.warning("No universes found in sheet; using built-in OTS list.")
        return _DEFAULT_UNIVERSES

    log.info(f"Loaded {len(names)} universes from sheet")
    return _build_universe_index(names)


def _word_variants(word: str) -> list:
    """
    Generate morphological variants of a word for fuzzy matching.
    Handles common English plural/suffix patterns.
    """
    w = word.lower()
    variants = [w]
    # Plurals → singular
    if w.endswith("ies") and len(w) > 4:
        variants.append(w[:-3] + "y")      # counties → county
    elif w.endswith("ves") and len(w) > 4:
        variants.append(w[:-3] + "f")      # leaves → leaf
    elif w.endswith("es") and len(w) > 4:
        variants.append(w[:-2])            # taxes → tax, nurses → nurs (close enough)
        variants.append(w[:-1])            # nurses → nurse
    elif w.endswith("s") and len(w) >= 4:
        variants.append(w[:-1])            # guns → gun, voters → voter
    # Singular → plural (for when keyword is singular but universe uses plural)
    if not w.endswith("s"):
        variants.append(w + "s")
    return list(dict.fromkeys(variants))   # deduplicate, preserve order


def _match_keyword_to_universe(word: str, universes: list) -> tuple:
    """
    Match a keyword to the best-fit voter targeting universe.

    Tries the word and its morphological variants (plural/singular forms).
    Scoring tiers (highest wins):
      0.90  — exact standalone word match in universe name tokens
      0.70  — substring match within the normalised universe name
      0.50  — meaningful prefix/root overlap between keyword and name token
    Only returns a match if score >= 0.50.

    Returns (universe_name, score_0_to_1) or (None, 0.0).
    """
    if not universes:
        return None, 0.0

    candidates = _word_variants(word)
    best_name  = None
    best_score = 0.0

    for u in universes:
        score = 0.0
        for wl in candidates:
            # Tier 1: exact word token match
            if wl in u["tokens"]:
                score = max(score, 0.90)
            # Tier 2: substring of normalised name
            elif wl in u["norm"]:
                score = max(score, 0.70)
            # Tier 3: meaningful prefix/root overlap
            else:
                if len(wl) >= 5:
                    for tok in u["tokens"]:
                        if len(tok) >= 5 and (tok.startswith(wl[:5]) or wl.startswith(tok[:5])):
                            score = max(score, 0.50)

        if score > best_score:
            best_score = score
            best_name  = u["name"]

    if best_score >= 0.50:
        return best_name, round(best_score, 2)
    return None, 0.0


# ---------------------------------------------------------------------------
# JSON export  (dashboard data file)
# ---------------------------------------------------------------------------

def _build_dashboard_json(articles: list[Article], universes: list) -> dict:
    """
    Build the JSON payload consumed by the GitHub Pages dashboard.

    Structure:
    {
      "meta":     { "generated_at": "...", "total": N },
      "articles": [ { title, author, date_published, body, url, source_site }, ... ],
      "keywords": [ { word, count, universe_match, universe_score }, ... ],
      "sources":  { "site.com": N, ... },
      "dates":    { "2026-04-09": N, ... }
    }
    """
    import json

    # Source counts
    sources: dict = {}
    dates:   dict = {}
    for a in articles:
        sources[a.source_site] = sources.get(a.source_site, 0) + 1
        if a.date_published:
            dates[a.date_published] = dates.get(a.date_published, 0) + 1

    # Keywords with universe matches
    raw_keywords = _extract_keywords(articles)
    keywords_out = []
    for word, count in raw_keywords:
        uni_name, uni_score = _match_keyword_to_universe(word, universes)
        keywords_out.append({
            "word":           word,
            "count":          count,
            "universe_match": uni_name,
            "universe_score": uni_score if uni_name else None,
        })

    # Articles — omit body from JSON to keep file size manageable;
    # store a short excerpt instead
    articles_out = []
    for a in articles:
        articles_out.append({
            "title":          a.title,
            "author":         a.author,
            "date_published": a.date_published,
            "body":           (a.body or "")[:500].strip(),   # excerpt only
            "url":            a.url,
            "source_site":    a.source_site,
        })

    return {
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "total":        len(articles),
        },
        "articles": articles_out,
        "keywords": keywords_out,
        "sources":  sources,
        "dates":    dates,
    }


def export_dashboard_json(
    articles: list[Article],
    json_path: str,
    universes_sheet: Optional[str] = None,
    append_mode: bool = False,
):
    """
    Write (or merge into) the dashboard JSON file.

    In append_mode, existing articles are preserved and new ones are merged in
    before re-computing keywords/sources/dates — so the dashboard always shows
    the full cumulative picture.
    """
    import json

    # Always load universes — uses built-in OTS list if no sheet specified
    universes = _load_universes(universes_sheet)

    if append_mode and os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            existing_urls = {a["url"] for a in existing.get("articles", [])}
            # Rebuild full Article objects from existing JSON so keyword
            # counting reflects the whole history, not just today's batch
            all_articles = []
            for a in existing.get("articles", []):
                all_articles.append(Article(
                    title=a.get("title") or "",
                    author=a.get("author"),
                    date_published=a.get("date_published"),
                    body=a.get("body") or "",
                    url=a.get("url") or "",
                    source_site=a.get("source_site") or "",
                ))
            # Append only genuinely new articles
            for a in articles:
                if a.url not in existing_urls:
                    all_articles.append(a)
            articles = all_articles
        except Exception as e:
            log.warning(f"Could not read existing JSON for merge ({e}); overwriting.")

    payload = _build_dashboard_json(articles, universes)

    os.makedirs(os.path.dirname(os.path.abspath(json_path)), exist_ok=True)
    import json as _json
    with open(json_path, "w", encoding="utf-8") as f:
        _json.dump(payload, f, ensure_ascii=False, indent=2)

    log.info(f"Wrote dashboard JSON → {json_path} "
             f"({len(payload['articles'])} articles, {len(payload['keywords'])} keywords)")


def git_push_dashboard(json_path: str):
    """
    Commit and push the updated data file to GitHub.

    Runs: git -C <repo_dir> add data/ && git commit -m "..." && git push
    """
    import subprocess
    repo_dir = os.path.dirname(os.path.abspath(json_path))
    # Walk up to find the git root
    check = repo_dir
    for _ in range(6):
        if os.path.isdir(os.path.join(check, ".git")):
            repo_dir = check
            break
        check = os.path.dirname(check)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    cmds  = [
        ["git", "-C", repo_dir, "add", "data/"],
        ["git", "-C", repo_dir, "commit", "-m", f"chore: update articles data {stamp}"],
        ["git", "-C", repo_dir, "push"],
    ]
    for cmd in cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            # "nothing to commit" is not a real error
            if "nothing to commit" in result.stdout + result.stderr:
                log.info("git: nothing to commit — data unchanged.")
                return
            log.warning(f"git command failed: {' '.join(cmd)}\n{result.stderr}")
            return
        log.info(f"git: {result.stdout.strip() or 'ok'}")
    print("  ✓ Dashboard data pushed to GitHub")


def main():
    args = parse_args()

    global MAX_ARTICLES_PER_SITE, DELAY_BETWEEN_REQUESTS, MAX_RETRIES
    MAX_ARTICLES_PER_SITE = args.max_articles
    DELAY_BETWEEN_REQUESTS = args.delay
    MAX_RETRIES = args.retries

    # Resolve the target date filter — build a set of YYYY-MM-DD strings
    from datetime import timedelta
    target_dates: Optional[set] = None
    if args.today:
        today = datetime.now().date()
        target_dates = {
            (today - timedelta(days=i)).strftime("%Y-%m-%d")
            for i in range(args.days)
        }
        label = ", ".join(sorted(target_dates))
        log.info(f"Date filter: last {args.days} day(s) — {label}")
    elif args.date:
        target_dates = {args.date}
        log.info(f"Date filter: {args.date}")

    # ------------------------------------------------------------------
    # Collect URLs — two modes:
    #   (a) Grouped mode: --sheet + --sheet-output-column
    #       → dict { output_filename: [url, ...] }
    #   (b) Flat mode: everything else
    #       → list of URLs, all written to args.output
    # ------------------------------------------------------------------

    use_grouped = bool(args.sheet and args.sheet_output_column)

    if use_grouped:
        # Grouped mode: read URL column + output-file column from the sheet
        url_groups = fetch_url_groups_from_google_sheet(
            args.sheet,
            url_column=args.sheet_column,
            output_column=args.sheet_output_column,
            tab_name=args.sheet_tab,
            default_output=args.output,
        )
        if not url_groups:
            print("No URLs found in the Google Sheet. Exiting.")
            sys.exit(1)

        total_sites = sum(len(v) for v in url_groups.values())
        print(f"\nScraping {total_sites} site(s) → {len(url_groups)} output file(s)...")
        for out_file, site_list in url_groups.items():
            print(f"  {out_file}: {len(site_list)} site(s)")
        print()

        fetcher = PageFetcher(use_selenium=args.selenium)
        grand_total = 0
        _grouped_articles_store = []   # accumulate all articles for JSON export
        try:
            for out_file, site_list in url_groups.items():
                group_articles = []
                for site_url in site_list:
                    if not site_url.startswith("http"):
                        site_url = "https://" + site_url
                    articles = scrape_site(
                        fetcher, site_url,
                        debug_html_path=args.debug_html,
                        debug_article_path=args.debug_article,
                        target_dates=target_dates,
                    )
                    group_articles.extend(articles)
                    log.info(f"  Collected {len(articles)} articles from {site_url}\n")

                if group_articles:
                    if target_dates and os.path.exists(out_file):
                        _append_to_xlsx(group_articles, out_file)
                    else:
                        write_xlsx(group_articles, out_file)
                    grand_total += len(group_articles)
                    _grouped_articles_store.append(group_articles)
                    print(f"  ✓ {len(group_articles)} articles → {out_file}")
                else:
                    print(f"  ✗ No articles extracted for {out_file}")
        finally:
            fetcher.close()

        print(f"\n{'='*50}")
        print(f"  Done! {grand_total} total articles saved across {len(url_groups)} file(s).")
        print(f"{'='*50}\n")

        # JSON dashboard export (all groups combined into one JSON)
        if args.export_json:
            all_grouped = [a for grp in _grouped_articles_store for a in grp]
            export_dashboard_json(
                all_grouped, args.export_json,
                universes_sheet=args.universes_sheet,
                append_mode=bool(target_dates),
            )
            if args.git_push:
                git_push_dashboard(args.export_json)

        return []  # groups were written individually

    # ------------------------------------------------------------------
    # Flat mode (original behaviour)
    # ------------------------------------------------------------------

    urls = []
    if args.urls:
        urls = args.urls
    elif args.sheet:
        urls = fetch_urls_from_google_sheet(
            args.sheet, tab_name=args.sheet_tab, column=args.sheet_column,
        )
    elif args.file:
        with open(args.file, "r") as f:
            urls = [line.strip() for line in f if line.strip() and not line.startswith("#")]
    else:
        urls = get_urls_interactive()

    if not urls:
        print("No URLs provided. Exiting.")
        sys.exit(1)

    # Normalize URLs
    urls = [u if u.startswith("http") else "https://" + u for u in urls]

    print(f"\nScraping {len(urls)} site(s)...")
    print(f"Output: {args.output}\n")

    # Initialize fetcher
    fetcher = PageFetcher(use_selenium=args.selenium)

    all_articles = []
    try:
        for site_url in urls:
            articles = scrape_site(
                fetcher, site_url,
                debug_html_path=args.debug_html,
                debug_article_path=args.debug_article,
                target_dates=target_dates,
            )
            all_articles.extend(articles)
            log.info(f"  Collected {len(articles)} articles from {site_url}\n")
    finally:
        fetcher.close()

    if all_articles:
        # In --today / --date mode, append to the xlsx so each day's run adds
        # new rows without overwriting previous days' articles.
        if target_dates and os.path.exists(args.output):
            _append_to_xlsx(all_articles, args.output)
        else:
            write_xlsx(all_articles, args.output)
        print(f"\n{'='*50}")
        print(f"  Done! {len(all_articles)} articles saved to {args.output}")
        print(f"{'='*50}\n")
    else:
        print("\nNo articles were extracted. This might happen if:")
        print("  - The site uses heavy JavaScript (try --selenium)")
        print("  - The site structure is unusual")
        print("  - The URLs were incorrect")
        print("  - The site blocks scraping requests\n")

    # ── JSON dashboard export ──────────────────────────────────────────────
    if args.export_json and all_articles:
        export_dashboard_json(
            all_articles, args.export_json,
            universes_sheet=args.universes_sheet,
            append_mode=bool(target_dates),
        )
        if args.git_push:
            git_push_dashboard(args.export_json)

    return all_articles


if __name__ == "__main__":
    main()
